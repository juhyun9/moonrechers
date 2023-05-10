
# -*- coding: utf-8 -*-

#필요한 모듈

from selenium import webdriver
import time
from openpyxl import Workbook
import pandas as pd
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import random
from selenium import webdriver
import chromedriver_autoinstaller
chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]  #크롬드라이버 버전 확인
options = webdriver.ChromeOptions()
# options.add_argument('--start-fullscreen')
# options.add_argument("headless")

import warnings 
warnings.filterwarnings('ignore')

youtuber_pd=pd.DataFrame()
comment_finals=[]
video_num= 2

#크롬드라이버 버전맞춰 자동 다운 및 시작
try:
    try:
        # 윈도우
        try:
            driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe', options=options)
        except:
            chromedriver_autoinstaller.install(True)
            driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe', options=options)
    except:
        # 맥
        try:
            driver = webdriver.Chrome(options=options)
        except:
            chromedriver_autoinstaller.install()
            driver = webdriver.Chrome(options=options)
except:
    print('크롬드라이버 자동설치 실패')


#그 유튜브 계정에 가서 관련 영상들 링크 크롤링
def takeContents(youtuber):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    driver = webdriver.Chrome("C:/Users\CPS/anaconda3/Scripts/chromedriver.exe")
    driver.get('https://www.youtube.com/'+youtuber+'/videos')
    driver.implicitly_wait(3)
    
    #인기순 정렬
    driver.find_element(By.XPATH,'//*[@id="chips"]/yt-chip-cloud-chip-renderer[2]').click()
    
    #웹페이지 끝까지 스크롤하기
    last_height = driver.execute_script("return document.documentElement.scrollHeight")

    while True:
        driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
        time.sleep(1.5)

        new_height = driver.execute_script("return document.documentElement.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    time.sleep(1.5)
    time.sleep(1.5)
    
    #비디오 링크 크롤링
    html_source = driver.page_source
    soup = BeautifulSoup(html_source, 'html.parser')
    global contentlist
    contentlist=[]
    global video_num
    number=0
    for href in soup.find_all("div", class_='style-scope ytd-rich-item-renderer'):
        number+=1
        hr= href.find("a")["href"]
        contentlist.append(hr)
        print(number)
        print(hr)
        takeComments(youtuber,hr,number)
        global video_num
        if number == video_num:
            break
    
    #contents_data = {'contents': contentlist}
    #Ycontents_pd = pd.DataFrame(contents_data)

#그 유튜브 계정에 가서 관련 영상들 링크 크롤링
def takeComments(youtuber,hr,number):
    url = 'https://www.youtube.com'+ str(hr)
    print(url)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    driver = webdriver.Chrome("C:/Users\CPS/anaconda3/Scripts/chromedriver.exe")
    driver.get(url)
    driver.implicitly_wait(3)
    time.sleep(1.5)

    driver.execute_script("window.scrollTo(0, 800)")
    time.sleep(3)
    
    #웹페이지 끝까지 스크롤하기
    last_height = driver.execute_script("return document.documentElement.scrollHeight")

    while True:
        driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
        time.sleep(1.5)

        new_height = driver.execute_script("return document.documentElement.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    time.sleep(1.5)
    
    #유튜브 팝업 닫기
    try:
        driver.find_element(By.CSS_SELECTOR, '#dismiss-button > a').click()
    except:
        pass
    
    #댓글 내용 크롤링
    html_source = driver.page_source
    soup = BeautifulSoup(html_source, 'html.parser')

    comment_list = soup.select("yt-formatted-string#content-text")
    comment_final=[]
    youtuber_final =[]

    for i in range(len(comment_list)):
        temp_comment = comment_list[i].text
        temp_comment = temp_comment.replace('\n', '')
        temp_comment = temp_comment.replace('\t', '')
        temp_comment = temp_comment.replace('    ', '')
        comment_final.append(temp_comment) # 댓글 내용
    
    todataframe(youtuber,number,comment_final)
    
    
def todataframe(youtuber,number, comment_final):
    for j in comment_final:
        global comment_finals
        comment_finals.append(j)
    if number == video_num:
        '''
        global df
        df={'youtuber'+str(number):youtuber,'comments':comment_finals}
        '''
        comm = ','.join(comment_finals)
        print(comm)
        print(youtuber)

        global youtuber_pd_pre
        youtuber_pd_pre = pd.DataFrame()
        youtuber_pd_pre["youtuber"] = [youtuber]
        youtuber_pd_pre["comments"] = [comm]

        global youtuber_pd
        youtuber_pd = pd.concat([youtuber_pd,youtuber_pd_pre])
        comment_finals=[]

def youtuberinput():
    inputyoutuber=str()
    while inputyoutuber != 'noone':
        inputyoutuber = input()
        if inputyoutuber not in youtuberlist :
            youtuberlistN.append(inputyoutuber)
            youtuberlist.append(inputyoutuber)
    youtuberlist.remove('noone')
    youtuberlistN.remove('noone')
    
#youtuberlist

#youtuberlist=['@shyfilms','@wonchan','@depublik']
youtuberrelist=['@EverydayIsThursdayNight', '@PsickUniv', '@jocoding', '@JTBCdrama', '@eo_studio', '@TheLocalProject', '@RISABAE', '@Bodeumofficial', '@MINEEEATS', '@samdaejang_official', '@beautyfool', '@mmtg_oops', '@youquizontheblock_official', '@GYMJONGKOOK', '@ootbstudio', '@Knowingbros', '@itsLiveOfficial', '@JBKWAK', '@ggondaehee', '@STUDIO_SUZE', '@EBS.LIFESTYLE', '@redpajamayaco', '@MBCtrueOn', '@Baeksang', '@tvNOfficial', '@rayejin', '@popfree', '@mrs_macarons', '@poong__e', '@jjaltoon', '@ITSUB', '@chongmmyung', '@little_sung_yup', '@TottenhamHotspur', '@officialbtob', '@VIVINOS', '@Ahn_19', '@alanbecker', '@JTBCMusic', '@tvNDENT', '@sungsikyung', '@shootforlovekorea', '@newsanhani', '@hunterfwang91', '@koreanenglishman', '@EatingwhatisGiven', '@somacguffin', '@odg.studio', '@user-nz8lo4cy8z', '@H___RYAN', '@kimhaejun', '@user-lb2kw3ge3p']
youtuberlistN=[]

for i in youtuberrelist:
    takeContents(i)

youtuber_pd.to_csv('crawling.csv', sep=',', encoding='utf-8')