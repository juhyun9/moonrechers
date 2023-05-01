# -*- coding: utf-8 -*-

#필요한 모듈

from selenium import webdriver
import time
from openpyxl import Workbook
import pandas as pd
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import random
from selenium import webdriver
import chromedriver_autoinstaller
chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]  #크롬드라이버 버전 확인
options = webdriver.ChromeOptions()
# options.add_argument('--start-fullscreen')
# options.add_argument("headless")

import warnings 
warnings.filterwarnings('ignore')

global youtuber_pd
youtuber_pd=pd.DataFrame()
global comment_finals
comment_finals=pd.DataFrame()
global comment_final
comment_final=[]
global pd_data
pd_data={}

#크롬드라이버 버전맞춰 자동 다운 및 시작
try:
    try:
        # 윈도우
        try:
            driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe', options=options)
        except:
            chromedriver_autoinstaller.install(True)
            driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe', options=options)
    except:
        # 맥
        try:
            driver = webdriver.Chrome(options=options)
        except:
            chromedriver_autoinstaller.install()
            driver = webdriver.Chrome(options=options)
except:
    print('크롬드라이버 자동설치 실패')

#2030유튜버 랭킹 순위 알 코드


#그 유튜브 계정에 가서 관련 영상들 링크 크롤링
def takeContents(youtuber):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    driver = webdriver.Chrome("C:/Users\CPS/anaconda3/Scripts/chromedriver.exe")
    driver.get('https://www.youtube.com/'+youtuber+'/videos')
    driver.implicitly_wait(3)
    
    #인기순 정렬
    driver.find_element(By.XPATH,'//*[@id="chips"]/yt-chip-cloud-chip-renderer[2]').click()
    
    #웹페이지 끝까지 스크롤하기
    last_height = driver.execute_script("return document.documentElement.scrollHeight")

    while True:
        driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
        time.sleep(1.5)

        new_height = driver.execute_script("return document.documentElement.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    time.sleep(1.5)
    time.sleep(1.5)
    
    #비디오 링크 크롤링
    html_source = driver.page_source
    soup = BeautifulSoup(html_source, 'html.parser')
    global contentlist
    contentlist=[]
    number=0
    for href in soup.find_all("div", class_='style-scope ytd-rich-item-renderer'):
        number+=1
        hr= href.find("a")["href"]
        contentlist.append(hr)
        print('link crawling done')
        takeComments(youtuber,hr,number)
        if number == 4:
            break
    
    #contents_data = {'contents': contentlist}
    #Ycontents_pd = pd.DataFrame(contents_data)
    
    
def takeComments(youtuber,content,number):
    #유튜브 불러오기,유튜버의 영상 1개
    url = 'https://www.youtube.com'+ str(content)
    print(url)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    driver = webdriver.Chrome("C:/Users\CPS/anaconda3/Scripts/chromedriver.exe")
    driver.get(url) #이 주소들 크롤링 선작업 필요
    driver.implicitly_wait(3)
    time.sleep(1.5)

    driver.execute_script("window.scrollTo(0, 800)")
    time.sleep(3)
    
    #웹페이지 끝까지 스크롤하기
    last_height = driver.execute_script("return document.documentElement.scrollHeight")

    while True:
        driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
        time.sleep(1.5)

        new_height = driver.execute_script("return document.documentElement.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    time.sleep(1.5)
    
    #유튜브 팝업 닫기
    try:
        driver.find_element(By.CSS_SELECTOR, '#dismiss-button > a').click()
    except:
        pass
    
    #댓글 내용 크롤링
    html_source = driver.page_source
    soup = BeautifulSoup(html_source, 'html.parser')

    comment_list = soup.select("yt-formatted-string#content-text")
    youtuber_final =[]

    for i in range(len(comment_list)):
        temp_comment = comment_list[i].text
        temp_comment = temp_comment.replace('\n', '')
        temp_comment = temp_comment.replace('\t', '')
        temp_comment = temp_comment.replace('    ', '')
        comment_final.append(temp_comment) # 댓글 내용

        

    
youtuberlist=['@EverydayIsThursdayNight','@PsickUniv','@jocoding']
for i in youtuberlist:
    takeContents(i)